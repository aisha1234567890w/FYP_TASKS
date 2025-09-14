import os, re, datetime, platform, difflib
import pyttsx3, speech_recognition as sr
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

EXCEL_FILE = "user_records.xlsx"
SHEET_NAME = "Details"

# ---------- TTS SETUP ----------
try:
    engine = pyttsx3.init("sapi5" if platform.system() == "Windows" else None)
    engine.setProperty("rate", 170)
    engine.setProperty("volume", 1.0)
except:
    engine = None

def speak(text):
    print(f"Assistant: {text}")
    if engine:
        try:
            engine.stop()
            engine.say(text)
            engine.runAndWait()
        except:
            pass

def listen_once(timeout=5, phrase_time_limit=7):
    r = sr.Recognizer()
    try:
        with sr.Microphone() as source:
            r.adjust_for_ambient_noise(source, duration=0.5)
            print("Listening...")
            audio = r.listen(source, timeout=timeout, phrase_time_limit=phrase_time_limit)
            return r.recognize_google(audio)
    except:
        return None

def ask(prompt):
    speak(prompt)
    resp = listen_once()
    if not resp:
        resp = input(f"{prompt}\nType your answer: ").strip()
    print(f"You: {resp}")
    return resp

def normalize_email(text):
    if not text: return ""
    s = text.lower().strip()
    s = re.sub(r"\bat sign\b|\bat the rate\b|\bat\b", "@", s)
    s = re.sub(r"\bdot com\b", ".com", s)
    s = re.sub(r"\bdot\b|\bperiod\b", ".", s)
    return s.replace(" ", "")

def normalize_phone(text):
    if not text: return ""
    s = text.lower()
    s = re.sub(r"\bplus\b", "+", s)
    w2d = {"zero":"0","oh":"0","one":"1","two":"2","three":"3","four":"4",
           "five":"5","six":"6","seven":"7","eight":"8","nine":"9"}
    def repl(m): return "".join(w2d.get(x, x) for x in m.group().split())
    s = re.sub(r"\b(?:zero|oh|one|two|three|four|five|six|seven|eight|nine)(?:\s+(?:zero|oh|one|two|three|four|five|six|seven|eight|nine))*\b", repl, s)
    digits = re.findall(r"\d+", s)
    return (s[0] if s.startswith("+") else "") + "".join(digits)

def get_gender():
    valid = {"male": "Male", "female": "Female"}
    while True:
        raw = ask("What is your gender?")
        heard = raw.lower().strip()
        if heard in valid:
            return valid[heard]
        match = difflib.get_close_matches(heard, valid.keys(), n=1, cutoff=0.6)
        if match:
            return valid[match[0]]
        speak("Please say male or female.")

def ensure_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        # --- Title Row ---
        ws.merge_cells("A1:E1")
        cell = ws.cell(row=1, column=1)
        cell.value = "User Details Records"
        cell.font = Font(size=16, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
        # --- Header Row ---
        headers = ["Timestamp", "Name", "Gender", "Email", "Phone"]
        for i, header in enumerate(headers, start=1):
            c = ws.cell(row=2, column=i)
            c.value = header
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(i)].width = 22
        wb.save(EXCEL_FILE)

def save_record(name, gender, email, phone):
    ensure_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([ts, name, gender, email, phone])
    wb.save(EXCEL_FILE)

def main():
    hour = datetime.datetime.now().hour
    greet = "Good morning!" if hour<12 else ("Good afternoon!" if hour<18 else "Good evening!")
    speak(f"{greet} Let's collect your details.")

    name = ask("What is your full name?")
    gender = get_gender()

    while True:
        raw_email = ask("Please tell me your email address.")
        email = normalize_email(raw_email)
        if re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
            break
        speak("That didn't look like a valid email. Let's try again.")

    while True:
        raw_phone = ask("What is your phone number? Include country code.")
        phone = normalize_phone(raw_phone)
        if len(phone.lstrip('+')) >= 7:
            break
        speak("That number seemed too short. Please repeat.")

    save_record(name, gender, email, phone)
    speak(f"Thanks {name}, your details have been saved successfully. Goodbye!")

if __name__ == "__main__":
    main()
